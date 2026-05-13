# ============================================================
# Excel Report Writer Service
# File: app/services/report_writer.py
#
# Generates formatted Excel reports using openpyxl with
# color-coded outcomes, multiple tabs, and summary metrics.
# ============================================================

import logging
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.call_session import CallSession, CallOutcome, Sentiment
from app.models.campaign import Campaign
from app.prompts.report_schema import (
    OUTCOME_COLORS,
    HEADER_STYLE,
    DATA_STYLE,
    REPORT_TABS,
    build_dynamic_schema,
)

logger = logging.getLogger(__name__)


class ReportWriter:
    """Generates Excel reports from call session data."""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def generate_report(
        self,
        campaign_id: int,
        filepath: str,
        include_tabs: list[str] | None = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
    ) -> dict:
        """
        Generate a full Excel report for a campaign.

        Args:
            campaign_id: Campaign to report on
            filepath: Output file path
            include_tabs: Which tabs to include (default: all)
            date_from: Optional start date filter
            date_to: Optional end date filter

        Returns:
            Summary statistics dict
        """
        # Load campaign
        campaign = await self.db.get(Campaign, campaign_id)
        if not campaign:
            raise ValueError(f"Campaign {campaign_id} not found")

        # Build dynamic schema based on campaign questions
        questions = campaign.questions or []
        schema = build_dynamic_schema(questions)

        # Load all call sessions
        query = select(CallSession).where(CallSession.campaign_id == campaign_id)
        if date_from:
            query = query.where(CallSession.created_at >= date_from)
        if date_to:
            query = query.where(CallSession.created_at <= date_to)
        query = query.order_by(CallSession.created_at.asc())

        result = await self.db.execute(query)
        sessions = result.scalars().all()

        # Convert sessions to row data
        rows = [self._session_to_row(s, schema) for s in sessions]

        # Create workbook
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        selected_tabs = include_tabs or list(REPORT_TABS.keys())
        stats = {}

        for tab_key in selected_tabs:
            tab_config = REPORT_TABS.get(tab_key)
            if not tab_config:
                continue

            if tab_key == "summary":
                # Special handling for summary tab
                self._write_summary_tab(wb, sessions, campaign)
                stats = self._calculate_stats(sessions)
            else:
                # Filter rows based on tab config
                filtered_rows = self._filter_rows(rows, tab_config.get("filter"))
                self._write_data_tab(
                    wb,
                    tab_name=tab_config["name"],
                    schema=schema,
                    rows=filtered_rows,
                )

        # Save workbook
        wb.save(filepath)
        logger.info(f"Report saved: {filepath} ({len(rows)} rows)")

        if not stats:
            stats = self._calculate_stats(sessions)

        return stats

    def _session_to_row(self, session: CallSession, schema: dict) -> dict:
        """Convert a CallSession to a row dict matching the schema."""
        extracted = session.extracted_data or {}
        answers = session.question_answers or {}

        row = {}
        for col, config in schema.items():
            json_key = config["json_key"]

            # Check extracted data first, then session fields
            if json_key in extracted:
                value = extracted[json_key]
            elif json_key in answers:
                value = answers[json_key]
            elif hasattr(session, json_key):
                attr = getattr(session, json_key)
                if hasattr(attr, "value"):  # Enum
                    value = attr.value
                else:
                    value = attr
            else:
                value = None

            # Format special types
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, bool):
                value = "Yes" if value else "No"
            elif value is None:
                value = ""

            row[col] = value

        # Store outcome for color coding
        row["_outcome"] = session.call_outcome.value if session.call_outcome else ""

        return row

    def _write_data_tab(
        self,
        wb: Workbook,
        tab_name: str,
        schema: dict,
        rows: list[dict],
    ):
        """Write a data tab with headers and rows."""
        ws = wb.create_sheet(title=tab_name)

        # --- Header Row ---
        header_fill = PatternFill(
            start_color=HEADER_STYLE["fill_color"],
            end_color=HEADER_STYLE["fill_color"],
            fill_type="solid",
        )
        header_font = Font(
            bold=HEADER_STYLE["font_bold"],
            color=HEADER_STYLE["font_color"],
            size=12,
            name="Calibri",
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, (col_letter, config) in enumerate(schema.items(), 1):
            cell = ws.cell(row=1, column=col_idx, value=config["header"])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        ws.row_dimensions[1].height = HEADER_STYLE["row_height"]

        # --- Data Rows ---
        for row_idx, row_data in enumerate(rows, 2):
            outcome = row_data.get("_outcome", "")
            row_color = OUTCOME_COLORS.get(outcome)

            for col_idx, col_letter in enumerate(schema.keys(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_letter, ""))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.font = Font(size=DATA_STYLE["font_size"], name="Calibri")

                if row_color:
                    cell.fill = PatternFill(
                        start_color=row_color,
                        end_color=row_color,
                        fill_type="solid",
                    )

            ws.row_dimensions[row_idx].height = DATA_STYLE["row_height"]

        # Auto-fit column widths (approximate)
        for col_idx, (col_letter, config) in enumerate(schema.items(), 1):
            header_len = len(config["header"])
            max_data_len = max(
                (len(str(row.get(col_letter, ""))) for row in rows),
                default=0,
            )
            width = min(max(header_len, max_data_len, 10) + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Apply auto-filter
        if rows:
            last_col = get_column_letter(len(schema))
            ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    def _write_summary_tab(self, wb: Workbook, sessions: list[CallSession], campaign: Campaign):
        """Write the summary statistics tab."""
        ws = wb.create_sheet(title="Summary")

        stats = self._calculate_stats(sessions)

        # --- Title ---
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")

        ws.merge_cells("A1:D1")
        title_cell = ws.cell(row=1, column=1, value=f"Campaign Report: {campaign.name}")
        title_cell.fill = header_fill
        title_cell.font = header_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        # --- Stats Grid ---
        stat_font = Font(size=12, name="Calibri")
        label_font = Font(bold=True, size=12, name="Calibri", color="1F4E79")

        stat_rows = [
            ("Campaign ID", campaign.campaign_id),
            ("Agent Name", campaign.agent_name),
            ("Company", campaign.company_name),
            ("Service", campaign.service_name),
            ("", ""),
            ("Total Calls Attempted", stats.get("total_calls_attempted", 0)),
            ("Total Calls Connected", stats.get("total_calls_connected", 0)),
            ("Connection Rate", f"{stats.get('connection_rate', 0):.1f}%"),
            ("Average Call Duration", f"{stats.get('average_call_duration', 0):.0f} seconds"),
            ("", ""),
            ("OUTCOME BREAKDOWN", ""),
            ("Interested", stats.get("outcome_breakdown", {}).get("interested", 0)),
            ("Not Interested", stats.get("outcome_breakdown", {}).get("not_interested", 0)),
            ("Callback Requested", stats.get("outcome_breakdown", {}).get("callback_requested", 0)),
            ("Voicemail", stats.get("outcome_breakdown", {}).get("voicemail", 0)),
            ("No Answer", stats.get("outcome_breakdown", {}).get("no_answer", 0)),
            ("Incomplete", stats.get("outcome_breakdown", {}).get("incomplete", 0)),
            ("", ""),
            ("SENTIMENT SPLIT", ""),
            ("Positive", stats.get("sentiment_split", {}).get("positive", 0)),
            ("Neutral", stats.get("sentiment_split", {}).get("neutral", 0)),
            ("Negative", stats.get("sentiment_split", {}).get("negative", 0)),
            ("", ""),
            ("Demo Requests", stats.get("demo_requests", 0)),
            ("Email Requests", stats.get("email_requests", 0)),
            ("", ""),
            ("Report Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
        ]

        for idx, (label, value) in enumerate(stat_rows, 3):
            label_cell = ws.cell(row=idx, column=1, value=label)
            label_cell.font = label_font
            value_cell = ws.cell(row=idx, column=2, value=value)
            value_cell.font = stat_font

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25

    def _filter_rows(self, rows: list[dict], filter_config) -> list[dict]:
        """Filter rows based on tab filter configuration."""
        if filter_config is None:
            return rows

        if not isinstance(filter_config, dict):
            return rows

        filtered = []
        for row in rows:
            match = True
            for key, value in filter_config.items():
                row_value = row.get("_outcome", "")
                if isinstance(value, list):
                    if row_value not in value:
                        match = False
                else:
                    if row_value != value:
                        match = False
            if match:
                filtered.append(row)

        return filtered

    def _calculate_stats(self, sessions: list[CallSession]) -> dict:
        """Calculate campaign-level statistics."""
        total = len(sessions)
        if total == 0:
            return {
                "total_calls_attempted": 0,
                "total_calls_connected": 0,
                "connection_rate": 0.0,
                "outcome_breakdown": {},
                "average_call_duration": 0.0,
                "sentiment_split": {},
                "demo_requests": 0,
                "email_requests": 0,
            }

        # Connected = any call that was actually answered
        connected = [s for s in sessions if s.duration_seconds and s.duration_seconds > 0]
        connection_rate = (len(connected) / total * 100) if total > 0 else 0

        # Outcome breakdown
        outcomes = {}
        for s in sessions:
            outcome = s.call_outcome.value if s.call_outcome else "incomplete"
            outcomes[outcome] = outcomes.get(outcome, 0) + 1

        # Average duration (of connected calls only)
        durations = [s.duration_seconds for s in connected if s.duration_seconds]
        avg_duration = sum(durations) / len(durations) if durations else 0

        # Sentiment
        sentiments = {}
        for s in sessions:
            sent = s.sentiment.value if s.sentiment else "neutral"
            sentiments[sent] = sentiments.get(sent, 0) + 1

        # Counts
        demo_requests = sum(1 for s in sessions if s.demo_requested)
        email_requests = sum(1 for s in sessions if s.email_requested)

        return {
            "total_calls_attempted": total,
            "total_calls_connected": len(connected),
            "connection_rate": connection_rate,
            "outcome_breakdown": outcomes,
            "average_call_duration": avg_duration,
            "sentiment_split": sentiments,
            "demo_requests": demo_requests,
            "email_requests": email_requests,
        }
